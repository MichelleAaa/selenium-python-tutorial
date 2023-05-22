import openpyxl

# openpyxl allows you to open excel workbooks and loop over their values.
def dataGenerator():
    wk = openpyxl.load_workbook('C:\\Users\\UPDATE\\PycharmProjects\\pythonProject\\EndToEndAutomation\\Book1.xlsx')
    sh = wk['Sheet1']
    r= sh.max_row #this fetches the number of rows.
    li = []
    for i in range (1, r + 1):
        li1 = []
        un = sh.cell(i, 1)
        up = sh.cell(i, 2)
        li1.insert(0, un.value)
        li1.insert(1, up.value)
        li.insert(i-1, li1)
    print(li)
    return li

# dataGenerator()